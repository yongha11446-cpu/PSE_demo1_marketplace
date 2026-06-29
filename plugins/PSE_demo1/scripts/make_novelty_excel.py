# -*- coding: utf-8 -*-
# /한계노벨티 결과(한계점 분류 + 노벨티 점수)를 엑셀(2시트)로 저장하는 글로벌 헬퍼.
# 사용법: python make_novelty_excel.py <입력JSON> <출력xlsx>
# 입력 JSON 형식:
# {
#   "topic": "주제명",
#   "weights": {"originality":0.4, "feasibility":0.3, "impact":0.3},  # (선택)
#   "limitations": [
#     {"kind":"공통|개선가능|개선불가","limitation":"...","papers":"03,07",
#      "overcome":"미해결|극복됨(출처)","difficulty":"상/중/하(이유)","note":"..."}, ...
#   ],
#   "novelties": [
#     {"novelty":"...","from_limit":"03","originality":5,"orig_note":"선행 없음(...)",
#      "feasibility":4,"feas_note":"...","impact":4,"impact_note":"...",
#      "score":4.5,"prior":"없음|유사(저자 연도)"}, ...
#   ]
# }
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 헤더별 열 너비(가독성) — 없는 헤더는 기본 18
WIDTHS = {
    "분류": 12, "한계점": 48, "관련 논문(번호)": 16, "극복여부": 28, "개선난이도": 26, "비고": 30,
    "노벨티": 50, "출발 한계(번호)": 16, "독창성(점)": 30, "구현(점)": 30, "임팩트(점)": 30,
    "종합점수": 12, "선행연구 유무(출처)": 30,
}

# 한계점 분류 정렬 우선순위 (공통 → 개선가능 → 개선불가)
KIND_ORDER = {"공통": 0, "개선가능": 1, "개선불가": 2}

def main():
    # 인자 확인
    if len(sys.argv) < 3:
        print("사용법: python make_novelty_excel.py <입력JSON> <출력xlsx>")
        sys.exit(1)
    in_json, out_xlsx = sys.argv[1], sys.argv[2]

    # 입력 JSON 읽기 (UTF-8)
    with open(in_json, encoding="utf-8") as f:
        data = json.load(f)

    limitations = data.get("limitations", [])    # 한계점 목록
    novelties = data.get("novelties", [])         # 노벨티 목록
    topic = data.get("topic", "")                 # 주제명

    # 정렬: 한계점은 분류 우선순위, 노벨티는 종합점수 내림차순
    limitations = sorted(limitations, key=lambda x: KIND_ORDER.get(x.get("kind", ""), 9))
    novelties = sorted(novelties, key=lambda x: x.get("score", 0), reverse=True)

    wb = Workbook()

    # ── 공통 스타일 ──
    head_font = Font(bold=True, color="FFFFFF")          # 헤더 글씨(흰색·볼드)
    head_fill = PatternFill("solid", fgColor="4F81BD")   # 헤더 배경(파랑)
    wrap = Alignment(wrap_text=True, vertical="top")     # 줄바꿈
    head_align = Alignment(wrap_text=True, vertical="center", horizontal="center")  # 헤더 가운데

    def write_sheet(ws, headers, keys, rows):
        # 헤더 행 작성
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = head_align
        # 데이터 행 작성
        for r, item in enumerate(rows, 2):
            for c, k in enumerate(keys, 1):
                ws.cell(row=r, column=c, value=item.get(k, "")).alignment = wrap
        # 열 너비(헤더별) + 헤더 고정
        for c, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(h, 18)
        ws.freeze_panes = "A2"

    # 시트1: 한계점
    ws1 = wb.active
    ws1.title = "한계점"
    write_sheet(
        ws1,
        ["분류", "한계점", "관련 논문(번호)", "극복여부", "개선난이도", "비고"],
        ["kind", "limitation", "papers", "overcome", "difficulty", "note"],
        limitations,
    )

    # 시트2: 노벨티 (독창성·구현·임팩트 점수와 근거를 한 칸에 합쳐 표기)
    ws2 = wb.create_sheet("노벨티")
    write_sheet(
        ws2,
        ["노벨티", "출발 한계(번호)", "독창성(점)", "구현(점)", "임팩트(점)", "종합점수", "선행연구 유무(출처)"],
        ["novelty", "from_limit", "orig_cell", "feas_cell", "impact_cell", "score", "prior"],
        # 점수+근거를 "점 — 근거" 형태로 합쳐 표준화
        [{
            "novelty": x.get("novelty", ""),
            "from_limit": x.get("from_limit", ""),
            "orig_cell": f'{x.get("originality","")} — {x.get("orig_note","")}'.strip(" —"),
            "feas_cell": f'{x.get("feasibility","")} — {x.get("feas_note","")}'.strip(" —"),
            "impact_cell": f'{x.get("impact","")} — {x.get("impact_note","")}'.strip(" —"),
            "score": x.get("score", ""),
            "prior": x.get("prior", ""),
        } for x in novelties],
    )

    # 파일 점유 등으로 저장 실패 시 _v2 새 이름으로
    try:
        wb.save(out_xlsx)
        saved = out_xlsx
    except PermissionError:
        saved = out_xlsx.rsplit(".", 1)[0] + "_v2.xlsx"
        wb.save(saved)

    print(f"저장 완료: {saved}  (주제: {topic}, 한계 {len(limitations)}개, 노벨티 {len(novelties)}개)")

if __name__ == "__main__":
    main()
