# -*- coding: utf-8 -*-
# 문헌 가치평가표 + 연구방향 3대안을 엑셀(2시트)로 저장하는 글로벌 헬퍼.
# 사용법: python make_lit_excel.py <입력JSON> <출력xlsx>
# 입력 JSON 형식:
# {
#   "topic": "주제명",
#   "papers": [
#     {"rank":1,"title":"...","authors":"...","year":2021,"journal":"...",
#      "grade_if":"SCI/IF 5.2","citations":120,"fit":"상(이유)","limitation":"...",
#      "novelty":"...","access":"가능|접근불가(업로드 필요)"}, ...
#   ],
#   "directions": [
#     {"rank":1,"name":"최선안","what":"...","why":"...","novelty":"...","risk":"..."}, ...
#   ]
# }
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 헤더별 열 너비(가독성) — 없는 헤더는 기본 18
WIDTHS = {
    "순위": 6, "공정 단계": 18, "제목": 60, "저자(연도)": 20, "저널(등급/IF)": 24,
    "인용수": 12, "적합도": 30, "한계점": 30, "노벨티 여지": 28, "접근성": 18,
    "이름": 28, "무엇을": 42, "왜": 30, "노벨티": 30, "리스크": 26, "조합 문헌(번호)": 36,
}

def main():
    # 인자 확인
    if len(sys.argv) < 3:
        print("사용법: python make_lit_excel.py <입력JSON> <출력xlsx>")
        sys.exit(1)
    in_json, out_xlsx = sys.argv[1], sys.argv[2]

    # 입력 JSON 읽기 (UTF-8)
    with open(in_json, encoding="utf-8") as f:
        data = json.load(f)

    papers = data.get("papers", [])          # 문헌 목록
    directions = data.get("directions", [])  # 연구방향 대안
    topic = data.get("topic", "")            # 주제명

    wb = Workbook()

    # ── 공통 스타일 ──
    head_font = Font(bold=True, color="FFFFFF")          # 헤더 글씨(흰색·볼드)
    head_fill = PatternFill("solid", fgColor="4F81BD")   # 헤더 배경(파랑)
    wrap = Alignment(wrap_text=True, vertical="top")     # 줄바꿈

    head_align = Alignment(wrap_text=True, vertical="center", horizontal="center")  # 헤더 가운데

    def write_sheet(ws, headers, keys, rows):
        # 헤더 행 작성
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = head_align
        # 데이터 행 작성
        for r, item in enumerate(rows, 2):
            for c, k in enumerate(keys, 1):
                ws.cell(row=r, column=c, value=item.get(k, "")).alignment = wrap
        # 열 너비(헤더별) + 헤더 고정
        for c, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(h, 18)
        ws.freeze_panes = "A2"

    # 시트1: 문헌평가  (stage 키가 있으면 '공정 단계' 열을 자동 포함)
    ws1 = wb.active
    ws1.title = "문헌평가"
    has_stage = any(p.get("stage") for p in papers)  # 단계 정보가 하나라도 있으면 열 추가
    base_headers = ["순위", "제목", "저자(연도)", "저널(등급/IF)", "인용수", "적합도", "한계점", "노벨티 여지", "접근성"]
    base_keys = ["rank", "title", "authors_year", "journal_grade", "citations", "fit", "limitation", "novelty", "access"]
    if has_stage:
        headers = ["순위", "공정 단계"] + base_headers[1:]
        keys = ["rank", "stage"] + base_keys[1:]
    else:
        headers, keys = base_headers, base_keys
    write_sheet(
        ws1, headers, keys,
        # papers의 키를 표준화 (authors+year, journal+grade_if 합치기)
        [{
            "rank": p.get("rank", ""),
            "stage": p.get("stage", ""),
            "title": p.get("title", ""),
            "authors_year": f'{p.get("authors","")} ({p.get("year","")})',
            "journal_grade": f'{p.get("journal","")} {p.get("grade_if","")}'.strip(),
            "citations": p.get("citations", ""),
            "fit": p.get("fit", ""),
            "limitation": p.get("limitation", ""),
            "novelty": p.get("novelty", ""),
            "access": p.get("access", ""),
        } for p in papers],
    )

    # 시트2: 연구방향 (어떤 문헌 번호를 조합했는지 열 포함)
    ws2 = wb.create_sheet("연구방향")
    write_sheet(
        ws2,
        ["순위", "이름", "조합 문헌(번호)", "무엇을", "왜", "노벨티", "리스크"],
        ["rank", "name", "papers", "what", "why", "novelty", "risk"],
        directions,
    )

    wb.save(out_xlsx)
    print(f"저장 완료: {out_xlsx}  (주제: {topic}, 문헌 {len(papers)}편, 방향 {len(directions)}개)")

if __name__ == "__main__":
    main()
