# -*- coding: utf-8 -*-
# 가상 시뮬레이션 결과를 엑셀(대안별 1시트 + 요약 1시트)로 저장하는 글로벌 헬퍼.
# 사용법: python make_vsim_excel.py <입력JSON> <출력xlsx>
# 입력 JSON 형식:
# {
#   "topic": "주제명",
#   "basis": "건조 Kenaf 1000 kg/hr (임의 basis)",
#   "alternatives": [
#     {"name":"대안1","설명":"...",
#      "blocks":[
#        {"block":"탈수","ref":"46.Ouyang2009 La-HZSM-5","temp":"확인 필요","pressure":"확인 필요",
#         "conv":"100%","sel":"에틸렌 97%","stream":"에틸렌 216.9","converge":"OK / 0%",
#         "evidence":"Ouyang 2009, Catal.Lett.132, LHSV0.5 (원문 유료→온도·압력 확인필요)"}, ...
#      ],
#      "summary":{"jet_kg_hr":210.8,"yield_feed":"21.1%","yield_etoh":"57.4%","판정":"수렴 OK"}}
#   ]
# }
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["블록", "기준논문(번호+이름)", "온도", "압력", "전환율/수율",
           "선택도", "주요 스트림(kg/hr)", "수렴/닫힘오차", "근거(저자·연도·Table/쪽)"]
KEYS = ["block", "ref", "temp", "pressure", "conv", "sel", "stream", "converge", "evidence"]
WIDTHS = [16, 34, 12, 12, 14, 16, 20, 16, 50]

SUM_HEADERS = ["대안", "구성(설명)", "jet 제품(kg/hr)", "수율/피드", "수율/EtOH", "판정"]
SUM_KEYS = ["name", "설명", "jet_kg_hr", "yield_feed", "yield_etoh", "판정"]
SUM_WIDTHS = [8, 60, 14, 12, 12, 28]

def main():
    if len(sys.argv) < 3:
        print("사용법: python make_vsim_excel.py <입력JSON> <출력xlsx>"); sys.exit(1)
    in_json, out_xlsx = sys.argv[1], sys.argv[2]
    with open(in_json, encoding="utf-8") as f:
        data = json.load(f)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4F81BD")
    head_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ── 요약 시트(맨 앞) ──
    ws = wb.active; ws.title = "요약"
    ws.cell(row=1, column=1, value=f"주제: {data.get('topic','')}   |   기준: {data.get('basis','')}")
    for c, h in enumerate(SUM_HEADERS, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = head_font; cell.fill = head_fill; cell.alignment = head_align
    for r, alt in enumerate(data.get("alternatives", []), 4):
        s = alt.get("summary", {})
        row = {"name": alt.get("name",""), "설명": alt.get("설명",""), **s}
        for c, k in enumerate(SUM_KEYS, 1):
            ws.cell(row=r, column=c, value=row.get(k, "")).alignment = wrap
    for c, w in enumerate(SUM_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── 대안별 시트 ──
    for alt in data.get("alternatives", []):
        ws = wb.create_sheet(alt.get("name", "대안"))
        ws.cell(row=1, column=1, value=alt.get("설명", ""))
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = head_font; cell.fill = head_fill; cell.alignment = head_align
        for r, blk in enumerate(alt.get("blocks", []), 3):
            for c, k in enumerate(KEYS, 1):
                ws.cell(row=r, column=c, value=blk.get(k, "")).alignment = wrap
        for c, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A3"

    try:
        wb.save(out_xlsx)
        print(f"저장 완료: {out_xlsx}")
    except PermissionError:
        alt = out_xlsx.replace(".xlsx", "_v2.xlsx")
        wb.save(alt)
        print(f"원본이 열려 있어 새 이름으로 저장: {alt}")

if __name__ == "__main__":
    main()
